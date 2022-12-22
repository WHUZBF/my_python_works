'''行列式的类'''
import re

import numpy.linalg

from fraction import Fraction as Frac
from fraction import multiply as mul
from fraction import plus_minus as pm
from fraction import reciprocal as rec
from fraction import transform
import squart
# import fractions # python标准库里面有分数就他妈离谱
import copy      #拷贝函数调用
import numpy as np   #终于向numpy低头了
import openpyxl as op

#一些使用重构的函数
def choose(matrix,mod):
    '''选择矩阵'''
    if mod == 'r':
        mat = matrix.data[:]
        return mat
    elif mod == 'c':
        mat = matrix.D_T[:]
        return mat
    else:
        print("Error:Invalid argument")
        return 0


def flag(num):
    '''判断一个数是不是0'''
    if isinstance(num, (int,float)) and num == 0:
        return 1
    elif not isinstance(num, (int,float)) and num.numerator == 0:
        return 1    #是零就返回1
    else:
        return 0


def count_zero(mat,r):
    '''确定矩阵某行第一个非零元的位置'''
    for index,ele in enumerate(mat.data[r]):
        if not flag(ele):
            return index
    return len(mat.data[r])  #全是零返回列数


# 注意类的命名使用驼峰命名法，也就是不使用下划线，每个单词首字母大写（python之禅）
class Matrix:
    def __init__(self, data=[], print_D_T = True, calculate=True):  # 数据为可选参数
        """
        使用define命令可以直接定义一个矩阵,命令格式为"define 矩阵名"
        您可以定义任意一个您喜欢的名字作为矩阵名,但尽可能相对简洁便于以后调用
        define 输入格式是行与行之间使用 ; 分隔,元素与元素之间使用 , 分隔,支持分数,浮点数输入，浮点数会自动转换为分数
        例如: 1,2,3;-7/8,0.5,6
        输出将会是:
        [  1    2   3 ]
        [-7/8  1/2  6 ]
        """

        self.order, self.c_num, self.r_num = 0, 0, 0  # 定义矩阵阶数
        if data:
            self.c_num = len(data[0])
            self.r_num = len(data)
            if self.r_num == self.c_num:
                self.order = self.r_num
        self.data = data  # 列表嵌套可以使用self.data[]来索引
        if not data:
            self.input_num()   # 没有传入参数时提示用户输入
        self.D_T = self.transpose()      # 自动传入实参self
        if calculate:
            # 对于元素为根式的矩阵而言暂不支持该类型计算
            self.rank = self.mat_rank()     # 矩阵的秩注意python为脚本语言，不需要提前编译好函数，要用的时候再跳转
        if print_D_T:
            self.output_data(self.D_T, '转置矩阵为:')


    def write_np(self):
        '''这个函数可以将矩阵写入numpy格式'''
        '''比较重要，所以放在前面'''
        array = []
        for data in self.data:
            #分数转化为小数
            np_data=[x if isinstance(x,(int,float)) else (x.numerator/x.denominator) for x in data]
            array.append(np_data)
        mat = np.array(array)
        return mat


    def updata(self):
        '''更新矩阵相关数据'''
        self.r_num = len(self.data)
        self.c_num = len(self.data[0])
        if self.r_num == self.c_num:
            self.order = self.r_num
        self.D_T = self.transpose()


    def output_data(self,data,msg='矩阵为:\n'):
        '''输出数据'''
        '''第二版打印方法，居中空格数据显示，更为清晰'''
        print(msg)
        #首先遍历二维数组寻找最大的字符长度
        max_len = 1
        for i in range(len(data)):
            for j in range(len(data[0])):
                if isinstance(data[i][j],(int,float)) and len(str(int(data[i][j]))) > max_len:
                    max_len = len(str(int(data[i][j])))
                elif (not isinstance(data[i][j],(int,float))) and len(data[i][j].str) > max_len:
                    max_len = len(data[i][j].str)

        for array in data:
            array = [num.str if not (isinstance(num,(int,float))) else str(int(num)) for num in array]    #列表解析处理分数变量的打印
            print("\t\t[",end='')
            for num_str in array:
                #num_str是一个字符串变量，可以使用方法center进行输出对齐
                print(num_str.center(max_len,' ')+'  ', end='')
            print("\b\b]")


    def input_num(self):
        '''输入矩阵的数据'''
        '''要注意提取之后都是字符类型，需要转换为int'''
        #global element   #添加的全局变量语句
        data_str = input("输入数据,行与行之间用分号连接,每个元素之间用逗号隔开:\n")
        if data_str[-1] != ';':
            data_str += ';'
        matches = re.findall('[^;]+;',data_str)   #提取每一行
        c_num = []
        for match in matches:
            elements = []
            for element in re.findall('[^,]*,',match.rstrip(';')+','):    #提取每一个数
                element=element.rstrip(',')  # 去掉逗号
                m = re.match("(.+)/(.+)",element)     #支持录入负数
                if m:
                     elements.append(Frac(int(m.group(1)),int(m.group(2))))    #一定要注意这里的类型转换
                else:
                    if '.' in str(element):
                        elements.append(transform(float(element))) #浮点数转化为分数
                    else:
                        elements.append(int(element))    # 不是输入的分数时直接转换为整数类型

            self.data.append(elements)
            c_num.append(len(elements))
        if len(set(c_num)) == 1:
            if c_num[0] == len(self.data):
                self.order = len(self.data)  # 表示出矩阵阶数
            self.r_num = len(self.data)
            self.c_num = c_num[0]
            self.output_data(self.data,'您输入的矩阵为:')
        else:
            print("输入错误请重新输入")
            self.data.clear()
            self.input_num()


    def revise_data(self):
        """
        如果发现矩阵中某个数据输入错误,可以考虑使用revise命令来更改矩阵数据,不必重新完整输入
        命令格式为 'revise 数据集‘ ,其中数据集是由 '行,列,新数据;' 这样的格式构成的,例如:
        'revise A 1,2,-7/8;3,5;9;'
        """
        new_datas = input("输入数据位置和新数据：")
        if new_datas[-1] != ';':
            new_datas += ';'
        new_datas = re.findall('[^;]+;',new_datas)
        for new_data in new_datas:
            match = re.match('(\d+),(\d+),(.+);',new_data)
            if match:
                match_frac = re.match("(.*)/(.*)",match.group(3))
                try:
                    if match_frac:
                        data = Frac(int(match_frac.group(1)),int(match_frac.group(2)))
                    else:
                        data = int(match.group(3))
                except ValueError:
                    print('>>>输入数据错误')
                else:
                    try:
                        self.data[int(match.group(1))-1][int(match.group(2))-1] = data
                    except IndexError:
                        print('>>>输入位置超过范围')
                    else:
                        self.output_data(self.data,msg = '新矩阵为:\n')
            else:
                print(">>>输入不合法")


    def transpose(self):
        """
        ***要使用transpose命令,您必须先进入calculate_mode***
        使用transpose可以直接转置一个矩阵并输出,支持使用save_as保留结果
        """

        data_T = self.data[:]  #创建列表副本
        D_T = []
        for c in range(len(data_T[0])):    #避免产生index错误
            array = []
            for r in range(len(data_T)):
                array.append(data_T[r][c])
            D_T.append(array)
        return D_T


    '''矩阵的一些运算函数'''
    #暂未支持分数运算
    def exchange(self,r_num_1,r_num_2,mod='r'):
        '''行列交换'''
        r_num_1 -= 1
        r_num_2 -= 1
        '''选择模式'''
        mat = choose(self,mod)
        if mat:
            try:
                '''计算部分'''
                mat[r_num_1], mat[r_num_2] = mat[r_num_2], mat[r_num_1]

                self.data = mat

                if mod == 'c':
                    self.data = self.transpose()

                self.D_T = self.transpose()
            except IndexError:
                print(">>>IndexError")
                return 0
            return 1


    def times(self,r_num,time,mod='r'):
        '''倍乘函数'''
        r_num -= 1
        mat = choose(self,mod)
        if mat :
            try:
                '''计算部分'''
                for index,data in enumerate(mat[r_num][:]):
                    '''注意改变data的值不改变列表本身'''
                    if isinstance(data,int) and isinstance(time,int):
                        data *= time     #支持分数形式
                    else:
                        data = mul(time,data)
                    mat[r_num][index] = data

                self.data = mat

                if mod == 'c':
                    self.data = self.transpose()

                self.D_T = self.transpose()
            except IndexError:
                print(">>>IndexError")
                return 0
            return 1


    def k_r(self,r_num_1,r_num_2,times,mod='r'):
        '''倍加行变换'''
        #把2k倍后加到1上面去
        r_num_1 -= 1
        r_num_2 -= 1
        mat = choose(self,mod)
        if mat :
            try:
                '''计算部分'''
                if mod == 'r':
                    x = self.c_num
                else:
                    x = self.r_num
                for index in range(0,x):
                    if isinstance(mat[r_num_1][index],int) and isinstance(times,int) and isinstance(mat[r_num_2][index],int):
                        mat[r_num_1][index] += times * mat[r_num_2][index]
                    else:
                        mat[r_num_1][index] = pm(mat[r_num_1][index],mul(times,mat[r_num_2][index]))     #支持分数形式

                if mod == 'c':
                    self.data = self.transpose()

                self.D_T = self.transpose()
            except IndexError:
                print(">>>IndexError")
                return 0
            return 1


    def simplify(self,mod = 0):
        """
        ***要使用simplify/simplest命令,您必须先进入calculate_mode***
        要将矩阵化简为行简化阶梯型请使用 'simple 待化简矩阵名称' 注意这个算法的结果和目前主流计算软件(maple,matlab)计算结果有些许不同但等价
        要将矩阵化简为最简行阶梯型(相抵标准型)请使用 'simplest 待化简矩阵名称'
        您同样可以使用save_as来保存计算结果
        """

        simple_mat = copy.deepcopy(self)
        '''采用高斯消元法进行'''
        '''直接采用循环框架'''
        def sort_mat(simple_mat,mat_r):     # 函数内部定义的函数就不能再在外部调用了
            '''对r行以后的矩阵进行冒泡排序'''
            for i in range(simple_mat.r_num-mat_r-1):
                # 冒泡排序进行n-1次循环
                for r in range(mat_r,simple_mat.r_num-i-1):
                    if count_zero(simple_mat,r) > count_zero(simple_mat,r+1):
                        simple_mat.exchange(r+1,r+2)  # 交换程序exchange中的索引要+1
        sort_mat(simple_mat,0)  # 先排一次序
        r=1
        while r<simple_mat.r_num:
            if count_zero(simple_mat,r) == self.c_num:  # 全为0元素后直接退出
                break
            elif count_zero(simple_mat,r) > count_zero(simple_mat,r-1):
                r+=1
            else:
                k = mul(rec(simple_mat.data[r-1][count_zero(simple_mat,r-1)]),simple_mat.data[r][count_zero(simple_mat,r)])
                k=mul(-1,k)
                simple_mat.k_r(r+1,r,k)
                sort_mat(simple_mat,r)  # 再次排序

        if mod == 1 and self.rank != 0:
            '''当mod选择为1时求行最简阶梯型'''
            # 先数一下应该化简几个1
            if self.rank == self.r_num:
                circle = self.r_num
            else:
                for r in range(1,simple_mat.r_num):
                    if count_zero(simple_mat,r) != count_zero(simple_mat,r-1)+1 or count_zero(simple_mat,r)==self.c_num:
                        circle = r          # 数出来了
                        break
            # 开始化简，注意第一行首个非零元的位置
            # 获取首个非零元的位置
            first_ele = count_zero(simple_mat,0)
            for x in range(circle):
                simple_mat.times(x+1,rec(simple_mat.data[x][x+first_ele]))  # 全部化为1
            for x in sorted(range(circle-1),reverse = True):
                i = 1
                for y in range(x+1,circle):
                    k = mul(simple_mat.data[x][first_ele+x+i],-1)
                    simple_mat.k_r(x+1,y+1,k)
                    i += 1
        return simple_mat


    '''求矩阵的一些参数'''
    def det(self,ans_print=True,promot_print=True):
        """
        ***要使用det命令,您必须先进入calculate_mode***
        本程序使用递归的方法求行列式的值,且所有运算转为分数运算
        'det 要求行列式的矩阵的名称' 为det命令的语法格式
        输出为一个数,无法使用save_as保存为矩阵
        """

        '''数据形式全部使用分数形式便于运算'''
        det_mat = copy.deepcopy(self)  # 创建矩阵实例对象的深拷贝
        det = Frac(0,1)   # 这个变量是在后面引用的
        if self.r_num != self.c_num and promot_print:
            print(">>>Error:非方阵,不可求行列式的值")
        elif self.order == 1:
            if isinstance(self.data[0][0],(float,int)):
                det = Frac(int(self.data[0][0]),1)
            else:
                det = self.data[0][0]
            if ans_print:
                print(f">>>ans = {det.str}")
        elif self.order == 2:
            det = pm(mul(self.data[0][0],self.data[1][1]),mul(self.data[0][1],self.data[1][0]),-1)
            if ans_print:
                print(f">>>ans = {det.str}")
        elif self.order > self.rank:
            det = Frac(0, 1)
            if ans_print:
                print(f">>>ans = {det.str}")
        else:
            '''高斯消元法求行列式'''
            def gauss(det_mat, const=1):   # const = 1行列式外的乘积因子
                value = det_mat.change_row()
                const = mul(const, value)
                first = det_mat.data[0][0]

                first_inverse = rec(first)   # 创建首项的倒数

                const = mul(first, const)
                if flag(const):
                    det = Frac(0,1)     # 写作分数形式,避免出错
                    return det

                det_mat.times(1,first_inverse)

                for index in range(1,det_mat.order):
                    if flag(det_mat.data[index][0]):
                        pass  # 避免0情况
                    else:
                        if isinstance(det_mat.data[index][0],(int,float)):
                            det_mat.data[index][0] = Frac(int(det_mat.data[index][0]),1)  #转换为分数进行计算
                        time = Frac(-det_mat.data[index][0].numerator,det_mat.data[index][0].denominator)   #注意减号
                        det_mat.k_r(index+1,1,time)    #消元


                # 重新写入数据
                new_det = []
                for r in range(1,det_mat.order):
                    array = []
                    for c in range(1,det_mat.order):
                        array.append(det_mat.data[r][c])
                    new_det.append(array)
                det_mat.data = new_det
                det_mat.order -= 1
                det_mat.updata()  # 更新数据

                if det_mat.order == 1:
                    return mul(det_mat.data[0][0], const)
                else:
                    return mul(gauss(det_mat), const)   # 递归求解
            det = gauss(det_mat)
            if ans_print:
                print(f">>>ans = {det.str}")
        return det


    def change_row(self):
        '''交换到首项非零行'''
        i=0
        if flag(self.data[0][0]):
            '''交换程序'''
            for index in range(1,self.order):
                if not flag(self.data[index][0]):
                    '''不是0就与首行交换'''
                    self.exchange(1,index+1)
                    return -1
                i+=1
            if i == self.order-1:
                '''全部是0直接行列式的值为0'''
                return 0

        return 1


    def trac(self):
        """
        ***要使用inv命令,您必须先进入calculate_mode***
        使用trac命令,您可以很方便的计算出矩阵的迹,结果将会输出到对话框中
        命令格式为 'trac 要计算迹的矩阵名称'
        """
        if self.order:
            trac = 0
            for i in range(self.order):
                trac = pm(self.data[i][i],trac)
            self.trac = trac
            return trac
        else:
            print(">>>Error非方阵不可求迹")
            return False



    def adjugate(self):
        """
        ***要使用adj命令,您必须先进入calculate_mode***
        伴随矩阵的运算量比较庞大,您可以使用adj命令去计算一个矩阵的伴随矩阵,本计算器在应对一般阶数时效率还是比较高
        您仍可以使用save_as命令来保存计算结果为矩阵形式
        求伴随命令的格式为 'adj 要计算伴随矩阵的矩阵名称'
        """

        if not self.order:
            print(">>>Error:非方阵不可求伴随矩阵")
            return 0
        adj_mat_data = copy.deepcopy(self.data)
        for i in range(self.order):
            for j in range(self.order):
                adj_mat_data[i][j] = cof(self.data, i, j, self.order)
        adj_mat = Matrix(adj_mat_data,False)
        return Matrix(adj_mat.D_T,False)


    def reverse(self,print_flag=1):
        """
        ***要使用inv命令,您必须先进入calculate_mode***
        使用inv命令,程序会调用reverse函数(显然更标准的说法应该是inverse matrix)计算矩阵的逆矩阵
        程序使用高斯消元法计算逆矩阵,更赞的是,计算器的所有计算都是以分数形式进行的,您不用担心精度丢失问题
        计算结果会直接以矩阵形式直观的输出到屏幕,如果您想保存计算结果为另一个矩阵可以使用save_as命令
        求逆运算命令格式为 'inv 求逆的矩阵的名称'
        """
        det = self.det(False,False)
        if self.r_num != self.c_num:
            print(">>>Error:矩阵不是方阵,不可逆")
            return 0
        elif flag(det):
            print(">>>Error:矩阵为奇异矩阵,不可逆")
            return 0
        else:
            # else 嵌套 if 表示前面都不满足且满足某条件
            if print_flag:
                print(">>>矩阵可逆!")
        '''矩阵求逆开始'''
        work_mat = copy.deepcopy(self)  #创建一个当前对象的深拷贝
        #一定不要忘记每次操作都要对单位矩阵也进行一次
        reverse_mat = identity_matrix(self.order) #创建一个同阶单位矩阵

        def change_mat_r(mat,reverse_mat,change_c):
            '''将矩阵切换到第一个不为0的行'''
            '''第C列进行检索'''
            change_c -= 1
            #这个变量之间列表的索引关系很容易搞混，原因是我之前自作聪明
            #可以举个例子来进行判断
            if flag(mat.data[change_c][change_c]):  #是零就得换一下
                for i in range(change_c+1,mat.order):
                    if not flag(mat.data[i][change_c]):
                        mat.exchange(i+1,change_c+1)
                        reverse_mat.exchange(i+1,change_c+1)
            else:
                return 0

        def gauss_reverse(mat,reverse_mat):
            '''高斯消元'''
            order = mat.order
            for c in range(order):
                #注意此时是正在进行第c+1行
                change_mat_r(work_mat,reverse_mat,c+1)
                #然后将第一个元素化为1
                frac = rec(work_mat.data[c][c])
                work_mat.times(c+1,frac)
                reverse_mat.times(c+1,frac)
                #用后面的行减去第一行
                for c_2 in range(c+1,order):
                    #注意这个时候的操作对象是c_2+1行
                    k = mul(-1,work_mat.data[c_2][c])
                    work_mat.k_r(c_2+1,c+1,k)
                    reverse_mat.k_r(c_2+1,c+1,k)
                #第一轮结束矩阵化为上三角

            #将矩阵化为单位矩阵
            for c_3 in range(2,order+1):
                #从下往上进行
                c_3 = -c_3    #操作的行的index
                #这个时候是在对order+c_3+1行进行操作
                #进行循环，也是倒数着来
                for x in range(1,-c_3):
                    K = [mul(num,-1) for num in work_mat.data[c_3][c_3+1:]]  #注意不能到-1因为左闭右开
                    work_mat.k_r(order+c_3+1,order-x+1,K[-x])   #先减去最后一排
                    reverse_mat.k_r(order+c_3+1,order-x+1,K[-x])
            return reverse_mat


        reverse_mat = gauss_reverse(work_mat,reverse_mat)
        return reverse_mat


    def mat_rank(self):
        """
        ***要使用rank命令,您必须先进入calculate_mode***
        使用 'rank 矩阵名' 命令,程序会直接打印矩阵的秩
        """

        work_mat = self.simplify()
        for r in range(work_mat.r_num):
            if count_zero(work_mat,r) == work_mat.c_num:
                return r
        return work_mat.r_num


    def LU(self):
        """
        ***要使用LU命令,您必须先进入calculate_mode***
        使用LU命令可以将矩阵做LU分解为上三角和下三角矩阵
        但很多情况下不能保证可以LU分解,所以实际上我们做的是PLU分解(这部分知识可以自行百度)
        命令格式为 'LU 待分解矩阵名称'
        """

        if self.r_num != self.c_num:
            print(">>>矩阵不是方阵不支持LU分解")
            return 0
        elif flag(self.det(False)):
            print(">>>矩阵为奇异矩阵不支持LU分解")
            return 0
        U = self.simplify(0)
        '''这里决定牺牲效率做完U后直接使用逆矩阵方法求L'''
        # L = identity_matrix(self.order)
        P_L = identity_matrix(self.order)   #这是要对L进行的行变换
        P = mat_mul(U,self.reverse(0))
        L = P.reverse(0)

        '''根据主对角元是不是1来进行排序'''
        for i in range(0,self.order):
            if not flag(pm(L.data[i][i],1,-1)):  # 检测是不是1
                for j in range(i+1,self.order):
                    if flag(pm(L.data[j][i],1,-1)):
                        L.exchange(i+1,j+1)
                        P_L.exchange(i+1,j+1)  # 做同样操作

        P_L.output_data(P_L.data, "P =")
        L.output_data(L.data, "L =")
        U.output_data(U.data, "U =")
        return (P,L,U)


    def schmidt(self):
        """
        ***要使用schmidt命令,您必须先进入calculate_mode***
        输入一个矩阵,程序会将其自动按列分割为一组列向量,然后使用这一组向量中的第一个(第一列)为起始向量,使用施密特正交化方法生成一组正交基
        程序会返回两个输出,一个是未经单位化的向量组,另一个是经过单位化的向量组,如果输入向量组线性相关,程序会报错并退出
        注意,使用save_as保存的是未经单位化的向量组的矩阵,您可以使用unit命令来保存单位化后的矩阵,结合value命令保留数值矩阵
        语法为 'schmidt 要进行正交化的向量组'
        """

        if self.rank != self.c_num:
            #检测是否列满秩
            print(">>>向量组中的向量线性相关,不能作为基底")
            return None
        sch_mat_data = []
        for i in range(0,self.c_num):
            vector = self.D_T[i]
            for j in range(0,i):
                prjvector = prj(self.D_T[i],sch_mat_data[j])    #函数内部调用函数没有定义先后限制
                vector=addv(vector,prjvector,-1)
            sch_mat_data.append(vector)
        sch_mat = Matrix(sch_mat_data,False)
        sch_mat = Matrix(sch_mat.D_T,False)
        sch_mat.output_data(sch_mat.data,"正交基为:")
        #单位化
        sch_mat.unit()
        return sch_mat


    def eig(self):
        """
        ***要使用eig命令,您必须先进入calculate_mode***
        输入 'eig 矩阵名',程序会计算矩阵的特征值和特征向量并输出
        注意,计算特征值和特征向量使用的是numpy科学计算库,特征值的输出是默认位数,特征向量的输出经过了润色,可以根据format输出
        输出中的'j'代表的是虚数单位,输出无法使用save_as保留为矩阵格式
        """

        if self.r_num != self.c_num:
            print("矩阵不是方阵，无法计算特征值")
            return None
        eig_mat = self.write_np()  #方法的引用要使用实例加点号的方式
        eigenvalue, featurevector = np.linalg.eig(eig_mat)
        return_value=[eigenvalue, featurevector]
        return return_value


    def unit(self):
        """
        ***要使用unit命令,您必须先进入calculate_mode***
        使用 'unit 矩阵名' 程序会将矩阵的每个列向量进行单位化并输出,其中数据的存储格式是符号根式模式
        sqrt()代表开方运算,i代表虚数单位,注意unit单位化后的矩阵使用save_as保存后是不能直接进行运算的,不然会报错
        但是可以进一步使用value命令化为数值矩阵后再对数值矩阵进行幂运算或者特征值运算
        本计算器对分数数据的处理相当好,但是无理数却无能为力,因为本产品没有采用浮点数计算模式,在牺牲效率下极大的保留了运算精度
        """

        matrix = copy.deepcopy(self.data)
        for c in range(self.c_num):
            array_sum = 0
            for r in range(self.r_num):
                if isinstance(matrix[r][c],(int,float)):
                    matrix[r][c] = Frac(int(matrix[r][c]),1)
                value = copy.deepcopy(matrix[r][c])
                value.power(2) # 前面新建一个临时变量，就不会改变matrix的值了,因为操作的元素是一个实例，必须要深拷贝
                array_sum = pm(array_sum,value)
            sqrt_sum = squart.sqrt_fraction(array_sum)
            sqrt_sum.rec()
            for r in range(self.r_num):
                factor = mul(matrix[r][c],sqrt_sum.factor)
                matrix[r][c] = squart.SqrtFraction(factor,sqrt_sum.sqrt)
        self.output_data(matrix, "单位化结果为：")
        return Matrix(matrix, False, False)


    def pow_mat(self,num):
        if (not isinstance(num,int)) or self.r_num != self.c_num:
            print("表达式非法！")
            return None
        else:
            ans = copy.deepcopy(self)
            for x in range(num):
                ans = mat_mul(ans,self)
        return ans


def list_mul(data_1,data_2):
    "关于行列向量的乘积"
    sum = 0
    for x in range(len(data_1)):
        '''当作分数来计算数据,更简洁'''
        '''但是内存消耗也更大'''
        ans = mul(data_1[x],data_2[x])
        sum = pm(ans,sum)
    return sum


def prj(v_1,v_2):
    '''向量v_1在v_2方向上的投影向量'''
    prj_vector=[]
    k=mul(list_mul(v_1,v_2),rec(list_mul(v_2,v_2)))
    for x in v_2:
        prj_vector.append(mul(k,x))
    return prj_vector


def addv(v_1,v_2,mod=1):
    '''向量加减法'''
    if len(v_1)!=len(v_2):
        return None
    v_3=[]
    for x in range(0,len(v_1)):
        v_3.append(pm(v_1[x],v_2[x],mod))
    return v_3


def merge(mat_A,mat_B,mod = ','):
    '''合并两个矩阵'''
    merge_data = copy.deepcopy(mat_A.data)   #深度copy
    if mod == ',':
        '''使用增广矩阵方法合并矩阵'''
        for index,r_B in enumerate(mat_B.data):
            for data_B in r_B:
                merge_data[index].append(data_B)
    elif mod == ';':
        '''使用竖直方法合并矩阵'''
        for r_B in mat_B.data:
            merge_data.append(r_B)
    merge_mat = Matrix(merge_data, False)
    return merge_mat


def mat_mul(mat_A,mat_B) :
    '''矩阵的乘法'''
    if mat_A.c_num != mat_B.r_num:
        print(">>>Error矩阵不可乘")
        return 0
    ans_data = []
    for v_A in mat_A.data:
        array = []  #临时存放每一行计算结果
        for v_B in mat_B.D_T:
            array.append(list_mul(v_A,v_B)) #计算开始
        ans_data.append(array)
    return Matrix(ans_data,False,False)   # 这里不计算rank节省内存


def mat_num(num,mat):
    '''矩阵的数乘'''
    ans_mat = copy.deepcopy(mat)
    for r in range(ans_mat.r_num):
        ans_mat.times(r+1,num)
    return ans_mat

def mat_plus_minus(mat_A,mat_B,mod = 1):
    '''矩阵的加减法'''
    if mat_A.r_num != mat_B.r_num and mat_A.c_num != mat_B.c_num:
        print(">>>Error:矩阵不同型,不可加减")
        return 0
    array = []
    for r in range(mat_A.r_num):
        array_r = []
        for c in range(mat_A.c_num):
            array_r.append(pm(mat_A.data[r][c],mat_B.data[r][c],mod))
        array.append(array_r)
    ans_mat = Matrix(array,False)
    return(ans_mat)


def identity_matrix(order):
    '''构造一个单位矩阵'''
    e_data=[]
    for i in range(order):
        array = [1 if x == i else 0 for x in range(order)] #列表解析
        e_data.append(array)
    E = Matrix(e_data, False)
    return E


def cof(data,r,c,order):
    '''求代数余子式'''
    #注意是对r+1,c+1进行
    cof_data = []
    for r_1 in range(order):
        if r_1 != r:
            array = []
            for c_1 in range(order):
                if c_1 != c:
                    array.append(data[r_1][c_1])
            cof_data.append(array)
    cof_mat = Matrix(cof_data,False)
    cof = mul(cof_mat.det(False),(-1)**(r+c))    # 注意是代数余子式所以要 乘以-1的i+j次方 长期不用易忘记
    return cof


def load_matrix():
    """
    \t使用load命令可以很方便的将xlsx工作表中的数据以矩阵形式提取出来,命令格式为 'load 输出的矩阵名'
    \t您可以任意定义输出的矩阵名称,文件名输错后会让您重新输入,注意,请将要导入的文件放在矩阵计算器根目录下
    \t输入正确的文件名称后,软件将会自动寻找可以调用的工作表名称,您可以单独决定导入哪个工作簿,同样您会有多次输入机会
    \t最后您需要输入导入数据的范围,给出左上角数的位置和右下角数的位置,软件便会自动划定边框然后载入矩阵
    \t输入格式为 '左上角行,左上角列,右下角行,右下角列'
    """

    filename = input("请输入要导入的文件名(包括扩展名):")
    try:
        work_book = op.load_workbook(filename)
    except Exception:
        print("文件名输入错误，请重新输入")
        return load_matrix()   # 递归调用
    print(f"检测到如下工作簿{work_book.sheetnames}")
    while True:
        sheet_name = input("输入要使用的工作簿：")
        if sheet_name in work_book.sheetnames:
            break
        else:
            print("未找到工作簿!")
    sheet_data = work_book[sheet_name]
    while True:
        cells = input("请输入开始单元格和结束单元格(主对角线关系)").split(',')
        cells = [int(cell) for cell in cells]
        if cells[0]<=cells[2] and cells[1]<=cells[3] and min(cells)>0:
            break
        else:
            print("输入错误!")
    mat_data = []
    for r in range(cells[0],cells[2]+1):
        row_data = []
        for c in range(cells[1], cells[3]+1):
            data = sheet_data.cell(row=r, column=c).value
            if data:
                if isinstance(data,float):
                    data = transform(data)
            else:
                data = 0
            row_data.append(data)
        mat_data.append(row_data)
    print("导入成功！")
    return Matrix(mat_data, False)


def zeros(r,c):
    """生成一个0数组"""
    zero = []
    for _ in range(r):
        zero_array = [0 for _ in range(c)]
        zero.append(zero_array)
    return zero


def dot_mul(mat_1,mat_2,mod):
    """求矩阵点乘"""
    work_mat1 = copy.deepcopy(mat_1.data)
    work_mat2 = copy.deepcopy(mat_2.data)
    work_mat3 = zeros(mat_1.r_num, mat_2.c_num)
    if mod == '/':
        for r in range(mat_2.r_num):
            for c in range(mat_2.c_num):
                work_mat2[r][c] = rec(work_mat2[r][c])
    for r in range(mat_1.r_num):
        for c in range(mat_1.c_num):
            work_mat3[r][c] = mul(work_mat1[r][c],work_mat2[r][c])
    return Matrix(work_mat3, False)


def dot_calculate(mat,num,mode):
    """矩阵和数的四则运算"""
    work_mat = copy.deepcopy(mat.data)
    data_mat = []
    if mode in ['.+', '.-']:
        if mode == '.-':
            mode = -1
        else:
            mode = 1
        for mat_r in work_mat:
            # 遍历列表中的元素，这个是二维的嵌套列表，mat_r出来后是一个列表
            mat_r = [pm(x, num, mode) for x in mat_r]
            data_mat.append(mat_r)
    elif mode == '.*':
        for mat_r in work_mat:
            mat_r = [mul(x, num) for x in mat_r]
            data_mat.append(mat_r)
    elif mode == '.^':
        if not isinstance(num, int):
            print(">>>指数必须是整数")
            return None
        else:
            for mat_r in work_mat:
                mat_r = [x ** num if isinstance(x, (int, float)) else x.power(num) for x in mat_r]
                data_mat.append(mat_r)
    return Matrix(data_mat, False)


class FloatMatrix(Matrix):
    def __init__(self,data):
        super().__init__([[1]], False, False)   # 空列表返回False
        self.data = self.data_init(data)
        self.mat = np.array(data)
        try:
            self.eigenvalue, self.featurevector = np.linalg.eig(self.mat)    # 使用self声明变量是一个属性，方便了之后的调用
            self.det = np.linalg.det(self.mat)
            self.inv_mat = np.linalg.inv(self.mat)
        except numpy.linalg.LinAlgError:
            print("...目前矩阵非方阵,无法计算各项属性")
            self.eigenvalue, self.featurevector, self.det, self.inv_mat = None, None, None, None

    def data_init(self,data):
        '''数据写入'''
        for r in range(len(data)):
            for c in range(len(data[0])):
                if isinstance(data[r][c],(float,int)):
                    data[r][c] = float(data[r][c])
                else:
                    data[r][c] = data[r][c].value
        return data

    def pow_mat(self,num):
        '''重写求幂方法'''
        ans = np.linalg.matrix_power(self.mat, num)
        ans_list = ans.tolist()
        return FloatMatrix(ans_list)

    def output_fdata(self,msg,format_len):
        '''输出矩阵的元素'''
        print(msg)
        np.set_printoptions(precision=format_len)
        print(self.mat)

    def output(self,format_len):
        '''输出'''
        self.output_fdata('矩阵为：',format_len)
        print("矩阵行列式值为：")
        print(f"{self.det}")
        print("矩阵的逆矩阵为:")
        print(f"{self.inv_mat}")
        print("矩阵特征值为：")
        print(f"{self.eigenvalue}")
        print("矩阵特征向量为：")
        print(f"{self.featurevector}")


class BoolMatrix:
    """布尔矩阵类"""
    # 由于数据类型不易兼容所以不继承父类
    def __init__(self, data=None):
        self.data = data
        self.size = None
        if not self.data:
            self.data_init()
        else:
            self.size = (len(self.data), len(self.data[0]))

    def data_init(self):
        """
        使用 bool 定义一个布尔矩阵
        您可以使用任何您喜欢的名称为其命名,但最好不要是keyword
        首先您需要给出矩阵的尺寸,输入格式为: 行数,列数
        然后您只需要直接按照顺序(行的排列顺序额)输入byte串就行了,软件会根据先前定好的尺寸来分割字符串
        注意:仅在bool模式下可用
        异常处理:将其它非0整数输入直接转换为其布尔值1,超过size会自动进行截断
        """
        size = input("请输入矩阵尺寸,格式为 行,列:")
        size = tuple([int(x) for x in size.rstrip().split(',')])
        byte = input("请直接输入byte串系统会自动分割,不足补零:")
        byte += '0' * (size[0]*size[1] - len(byte))
        data_str = [byte[x*size[1]:(x+1)*size[1]] for x in range(size[0])]
        self.data = [[int(bool(int(j))) for j in list(i)] for i in data_str]
        self.size = size
        self.output_data(self.data, "...您输入的矩阵为:")

    def output_data(self, data, msg):
        print(msg)
        for r in range(self.size[0]):
            print("\t[", end='')
            for x in data[r]:
                print(x, end=" ")
            print("\b]")

    def NOT(self):
        """
        使用 not 或者 !来对矩阵取非
        注意:仅在bool模式下可用
        """
        not_data = [[int(not j) for j in i] for i in self.data]
        return BoolMatrix(not_data)

    def AND(self, mat):
        """
        使用and 或者 & 符号来对两个矩阵取交
        注意:仅在bool 模式下可用
        """
        # 程序语言天生适合逻辑运算
        if self.size != mat.size:
            print("...矩阵不同型!")
            return None
        mat_a = self.data
        mat_b = mat.data
        and_data = [[int(mat_a[r][c] and mat_b[r][c]) for c in range(self.size[1])] for r in range(self.size[0])]
        return BoolMatrix(and_data)

    def OR(self, mat):
        """
        使用 or 或者 | 符号来对两个矩阵取并
        注意:仅在 bool 模式下可用
        """
        if self.size != mat.size:
            print("...矩阵不同型!")
            return None
        mat_a = self.data
        mat_b = mat.data
        or_data = [[int(mat_a[r][c] or mat_b[r][c]) for c in range(self.size[1])] for r in range(self.size[0])]
        return BoolMatrix(or_data)

    def XOR(self, mat):
        """
        使用 xor 或者 ^ 符号来对两个矩阵取异或
        注意:仅在 bool 模式下可用,calculate_mode下会自动将^理解为矩阵的幂
        """
        if self.size != mat.size:
            print("...矩阵不同型!")
            return None
        mat_a = self.data
        mat_b = mat.data
        xor_data = [[int(mat_a[r][c] ^ mat_b[r][c]) for c in range(self.size[1])] for r in range(self.size[0])]
        return BoolMatrix(xor_data)

    def turn(self):
        T_data = []
        for c in range(self.size[1]):
            array = []
            for r in range(self.size[0]):
                array.append(self.data[r][c])
            T_data.append(array)
        return BoolMatrix(T_data)

    def BoolProduct(self, mat):
        """
        使用 # 来对矩阵求布尔积
        注意:仅在 bool 模式下可用
        """
        if self.size[1] != mat.size[0]:
            print("...矩阵不可乘!")
            return None
        mat_a = self.data
        mat_b = mat.turn().data
        product_mat = []

        def array_bool_product(a_1, a_2):
            product = a_1[0] and a_2[0]
            for i in range(1, len(a_1)):
                product = product or (a_1[i] and a_2[i])
            return product

        for v_a in mat_a:
            products = []
            for v_b in mat_b:
                products.append(array_bool_product(v_a, v_b))
            product_mat.append(products)
        return BoolMatrix(product_mat)

    def BoolPower(self, n):
        """
        使用 #^ 来对矩阵求布尔幂
        注意:仅在 bool 模式下可用
        """
        power_mat = self
        if (not isinstance(n, int)) or n<=0:
            print("...输入非法!")
            return None
        elif self.size[0] != self.size[1]:
            print("...不是方阵,不可求幂!")
            return None
        for _ in range(n-1):
            # 第一个实参自动传入是power_mat 第二个实参为self表示调用这个函数的矩阵,因为无法直接确定实例名称,所以使用self去表示调用这个方法的
            # 实例本身,是可以当作参数直接传入函数的,而且方法一定会自动传入实例本身,不管需不需要调用,没有调用self时便称为静态函数
            # 注意在外部不能使用self这个关键字,只能直接输入实例名代表实例本身
            power_mat = power_mat.BoolProduct(self)
        return power_mat